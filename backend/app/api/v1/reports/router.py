import io
import uuid
from datetime import date, datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependencies import CurrentUser
from app.db.session import get_db

router = APIRouter()


@router.get("/attendance")
async def attendance_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    format: str = Query("json"),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
):
    from app.models.employee import Attendance, Employee
    from calendar import monthrange

    _, days_in_month = monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    emp_result = await db.execute(select(Employee).where(Employee.is_active == True))
    employees = emp_result.scalars().all()

    att_result = await db.execute(
        select(Attendance).where(Attendance.date >= start, Attendance.date <= end)
    )
    attendances = att_result.scalars().all()

    att_map: dict[uuid.UUID, list[Attendance]] = {}
    for a in attendances:
        att_map.setdefault(a.employee_id, []).append(a)

    rows = []
    for emp in employees:
        records = att_map.get(emp.id, [])
        present = sum(1 for r in records if r.status == "present")
        half_day = sum(1 for r in records if r.status == "half_day")
        wfh = sum(1 for r in records if r.status == "wfh")
        absent = days_in_month - len(records)
        rows.append({
            "employee_id": str(emp.id),
            "employee_name": emp.name,
            "designation": emp.designation or "",
            "present": present,
            "half_day": half_day,
            "wfh": wfh,
            "absent": absent,
            "total_days": days_in_month,
        })

    if format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance {month}/{year}"
        headers = ["Employee", "Designation", "Present", "Half Day", "WFH", "Absent", "Total Days"]
        ws.append(headers)
        for r in rows:
            ws.append([r["employee_name"], r["designation"], r["present"], r["half_day"], r["wfh"], r["absent"], r["total_days"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=attendance_{year}_{month:02d}.xlsx"},
        )

    return {"month": month, "year": year, "rows": rows}


@router.get("/gst")
async def gst_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    format: str = Query("json"),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
):
    from app.models.finance import Invoice
    from calendar import monthrange

    _, days_in_month = monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    result = await db.execute(
        select(Invoice).where(
            Invoice.is_gst == True,
            Invoice.is_deleted == False,
            Invoice.invoice_date >= start,
            Invoice.invoice_date <= end,
        ).order_by(Invoice.invoice_date.asc())
    )
    invoices = result.scalars().all()

    rows = [{
        "invoice_number": inv.invoice_number,
        "invoice_date": inv.invoice_date.isoformat(),
        "customer_name": inv.customer_name,
        "customer_gstin": inv.customer_gstin or "",
        "taxable_value": float(inv.subtotal),
        "cgst": float(inv.cgst),
        "sgst": float(inv.sgst),
        "igst": float(inv.igst),
        "total": float(inv.total),
    } for inv in invoices]

    totals = {
        "taxable_value": sum(r["taxable_value"] for r in rows),
        "cgst": sum(r["cgst"] for r in rows),
        "sgst": sum(r["sgst"] for r in rows),
        "igst": sum(r["igst"] for r in rows),
        "total": sum(r["total"] for r in rows),
    }

    if format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"GST {month}/{year}"
        ws.append(["Invoice No.", "Date", "Customer", "GSTIN", "Taxable Value", "CGST", "SGST", "IGST", "Total"])
        for r in rows:
            ws.append([r["invoice_number"], r["invoice_date"], r["customer_name"], r["customer_gstin"],
                       r["taxable_value"], r["cgst"], r["sgst"], r["igst"], r["total"]])
        ws.append(["TOTAL", "", "", "", totals["taxable_value"], totals["cgst"], totals["sgst"], totals["igst"], totals["total"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=gst_{year}_{month:02d}.xlsx"},
        )

    return {"month": month, "year": year, "rows": rows, "totals": totals}


@router.get("/pl")
async def pl_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    format: str = Query("json"),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
):
    from app.models.finance import AccountEntry
    from calendar import monthrange

    _, days_in_month = monthrange(year, month)
    start = date(year, month, 1)
    end = date(year, month, days_in_month)

    result = await db.execute(
        select(AccountEntry).where(AccountEntry.date >= start, AccountEntry.date <= end)
    )
    entries = result.scalars().all()

    income: dict[str, float] = {}
    expense: dict[str, float] = {}
    for e in entries:
        if e.type == "income":
            income[e.category] = income.get(e.category, 0) + float(e.amount)
        else:
            expense[e.category] = expense.get(e.category, 0) + float(e.amount)

    income_rows = [{"category": k, "amount": v} for k, v in sorted(income.items())]
    expense_rows = [{"category": k, "amount": v} for k, v in sorted(expense.items())]
    total_income = sum(income.values())
    total_expense = sum(expense.values())

    if format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"P&L {month}/{year}"
        ws.append(["Type", "Category", "Amount"])
        for r in income_rows:
            ws.append(["Income", r["category"], r["amount"]])
        for r in expense_rows:
            ws.append(["Expense", r["category"], r["amount"]])
        ws.append(["", "Total Income", total_income])
        ws.append(["", "Total Expense", total_expense])
        ws.append(["", "Net Profit", total_income - total_expense])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=pl_{year}_{month:02d}.xlsx"},
        )

    return {
        "month": month, "year": year,
        "income": income_rows,
        "expense": expense_rows,
        "total_income": total_income,
        "total_expense": total_expense,
        "net": total_income - total_expense,
    }


@router.get("/tasks")
async def task_report(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2020),
    format: str = Query("json"),
    db: AsyncSession = Depends(get_db),
    current_user: CurrentUser = None,
):
    from app.models.task import Task
    from app.models.employee import Employee
    from calendar import monthrange
    from sqlalchemy.orm import selectinload

    _, days_in_month = monthrange(year, month)
    start = datetime(year, month, 1, tzinfo=timezone.utc)
    end = datetime(year, month, days_in_month, 23, 59, 59, tzinfo=timezone.utc)

    task_result = await db.execute(
        select(Task)
        .options(selectinload(Task.assigned_to))
        .where(Task.is_deleted == False, Task.due_date >= start, Task.due_date <= end)
    )
    tasks = task_result.scalars().all()

    emp_stats: dict[str, dict] = {}
    for t in tasks:
        name = t.assigned_to.name if t.assigned_to else "Unassigned"
        emp_id = str(t.assigned_to_id) if t.assigned_to_id else "unassigned"
        if emp_id not in emp_stats:
            emp_stats[emp_id] = {"name": name, "total": 0, "completed": 0, "overdue": 0, "completion_secs": []}
        emp_stats[emp_id]["total"] += 1
        if t.status == "completed":
            emp_stats[emp_id]["completed"] += 1
            if t.completed_at and t.created_at:
                delta = (t.completed_at - t.created_at).total_seconds()
                emp_stats[emp_id]["completion_secs"].append(delta)
        elif t.status == "overdue":
            emp_stats[emp_id]["overdue"] += 1

    rows = []
    for emp_id, s in emp_stats.items():
        secs = s["completion_secs"]
        avg_hours = round(sum(secs) / len(secs) / 3600, 1) if secs else None
        rate = round(s["completed"] / s["total"] * 100, 1) if s["total"] else 0
        rows.append({
            "employee_name": s["name"],
            "total": s["total"],
            "completed": s["completed"],
            "overdue": s["overdue"],
            "avg_completion_hours": avg_hours,
            "completion_rate": rate,
        })

    rows.sort(key=lambda r: -r["completion_rate"])

    if format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Task Report {month}/{year}"
        ws.append(["Employee", "Total Tasks", "Completed", "Overdue", "Avg Hours", "Completion %"])
        for r in rows:
            ws.append([r["employee_name"], r["total"], r["completed"], r["overdue"], r["avg_completion_hours"] or "", r["completion_rate"]])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=tasks_{year}_{month:02d}.xlsx"},
        )

    return {"month": month, "year": year, "rows": rows}
