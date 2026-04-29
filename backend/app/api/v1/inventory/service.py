import io
import uuid
from datetime import date
from typing import Optional

from fastapi import HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.v1.schemas import PaginatedResponse
from .model import Inventory, PlatformOrder, Product, ProductionBatch, RawMaterial
from .schema import (
    BatchCreate,
    BatchResponse,
    PlatformOrderCreate,
    PlatformOrderResponse,
    ProductCreate,
    ProductResponse,
    RawMaterialCreate,
    RawMaterialResponse,
    StockAdjust,
)


def _product_response(p: Product) -> ProductResponse:
    inv = p.inventory
    return ProductResponse(
        id=p.id,
        name=p.name,
        sku=p.sku,
        category=p.category,
        description=p.description,
        unit=p.unit,
        mrp=float(p.mrp),
        cost_price=float(p.cost_price),
        gst_rate=float(p.gst_rate),
        hsn_code=p.hsn_code,
        is_active=p.is_active,
        image_url=p.image_url,
        current_stock=float(inv.current_stock) if inv else 0,
        reserved_stock=float(inv.reserved_stock) if inv else 0,
        reorder_level=float(inv.reorder_level) if inv else 10,
        max_stock=float(inv.max_stock) if inv else 1000,
        is_low_stock=(float(inv.current_stock) <= float(inv.reorder_level)) if inv else False,
    )


async def list_products(
    db: AsyncSession,
    category: Optional[str] = None,
    low_stock: Optional[bool] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
) -> PaginatedResponse[ProductResponse]:
    query = select(Product).options(selectinload(Product.inventory)).where(Product.is_active == True)
    if category:
        query = query.where(Product.category == category)
    if search:
        term = f"%{search}%"
        query = query.where(or_(Product.name.ilike(term), Product.sku.ilike(term)))

    if low_stock:
        # fetch all matching, filter in Python (inventory join needed)
        result = await db.execute(query.order_by(Product.name))
        products = [
            p for p in result.scalars().all()
            if p.inventory and float(p.inventory.current_stock) <= float(p.inventory.reorder_level)
        ]
        total = len(products)
        paged = products[(page - 1) * limit : page * limit]
        return PaginatedResponse.build([_product_response(p) for p in paged], total, page, limit)

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()

    query = query.order_by(Product.name).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    items = [_product_response(p) for p in result.scalars()]
    return PaginatedResponse.build(items, total, page, limit)


async def create_product(
    db: AsyncSession, body: ProductCreate, actor_id: uuid.UUID, actor_name: str
) -> ProductResponse:
    from app.utils.activity import log as activity_log

    existing = await db.execute(select(Product).where(Product.sku == body.sku))
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="SKU already exists")

    product = Product(
        name=body.name,
        sku=body.sku,
        category=body.category,
        description=body.description,
        unit=body.unit,
        mrp=body.mrp,
        cost_price=body.cost_price,
        gst_rate=body.gst_rate,
        hsn_code=body.hsn_code,
    )
    db.add(product)
    await db.flush()

    inv = Inventory(
        product_id=product.id,
        current_stock=0,
        reorder_level=body.reorder_level,
        max_stock=body.max_stock,
    )
    db.add(inv)
    await db.flush()

    await activity_log(
        db,
        actor_id=actor_id,
        actor_name=actor_name,
        action="create",
        entity_type="product",
        entity_id=str(product.id),
        entity_name=product.name,
        description=f"{actor_name} added product '{product.name}' (SKU: {product.sku})",
    )

    product.inventory = inv
    return _product_response(product)


async def get_product(db: AsyncSession, product_id: str) -> ProductResponse:
    result = await db.execute(
        select(Product).options(selectinload(Product.inventory)).where(Product.id == uuid.UUID(product_id))
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return _product_response(product)


async def adjust_stock(
    db: AsyncSession, product_id: str, body: StockAdjust, actor_id: uuid.UUID, actor_name: str
) -> ProductResponse:
    from app.utils.activity import log as activity_log

    result = await db.execute(
        select(Product).options(selectinload(Product.inventory)).where(Product.id == uuid.UUID(product_id))
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")

    if not product.inventory:
        inv = Inventory(product_id=product.id, current_stock=body.quantity)
        db.add(inv)
        product.inventory = inv
    else:
        new_stock = float(product.inventory.current_stock) + body.quantity
        if new_stock < 0:
            raise HTTPException(status_code=400, detail="Insufficient stock")
        product.inventory.current_stock = new_stock

    await db.flush()

    direction = "added" if body.quantity >= 0 else "removed"
    await activity_log(
        db,
        actor_id=actor_id,
        actor_name=actor_name,
        action="update",
        entity_type="product",
        entity_id=str(product.id),
        entity_name=product.name,
        description=f"{actor_name} {direction} {abs(body.quantity)} {product.unit} stock for '{product.name}' — {body.reason}",
    )

    return _product_response(product)


async def list_raw_materials(db: AsyncSession) -> list[RawMaterialResponse]:
    result = await db.execute(select(RawMaterial).order_by(RawMaterial.name))
    materials = result.scalars().all()
    return [
        RawMaterialResponse(
            id=m.id,
            name=m.name,
            unit=m.unit,
            current_stock=float(m.current_stock),
            reorder_level=float(m.reorder_level),
            cost_per_unit=float(m.cost_per_unit),
            supplier=m.supplier,
            notes=m.notes,
            is_low_stock=float(m.current_stock) <= float(m.reorder_level),
        )
        for m in materials
    ]


async def create_raw_material(db: AsyncSession, body: RawMaterialCreate) -> RawMaterialResponse:
    mat = RawMaterial(**body.model_dump())
    db.add(mat)
    await db.flush()
    return RawMaterialResponse(
        id=mat.id,
        name=mat.name,
        unit=mat.unit,
        current_stock=float(mat.current_stock),
        reorder_level=float(mat.reorder_level),
        cost_per_unit=float(mat.cost_per_unit),
        supplier=mat.supplier,
        notes=mat.notes,
        is_low_stock=float(mat.current_stock) <= float(mat.reorder_level),
    )


async def adjust_raw_material(db: AsyncSession, material_id: str, body: StockAdjust) -> RawMaterialResponse:
    result = await db.execute(select(RawMaterial).where(RawMaterial.id == uuid.UUID(material_id)))
    mat = result.scalar_one_or_none()
    if not mat:
        raise HTTPException(status_code=404, detail="Material not found")
    mat.current_stock = float(mat.current_stock) + body.quantity
    return RawMaterialResponse(
        id=mat.id,
        name=mat.name,
        unit=mat.unit,
        current_stock=float(mat.current_stock),
        reorder_level=float(mat.reorder_level),
        cost_per_unit=float(mat.cost_per_unit),
        supplier=mat.supplier,
        notes=mat.notes,
        is_low_stock=float(mat.current_stock) <= float(mat.reorder_level),
    )


async def list_batches(
    db: AsyncSession, status: Optional[str] = None
) -> list[BatchResponse]:
    query = select(ProductionBatch)
    if status:
        query = query.where(ProductionBatch.status == status)
    query = query.order_by(ProductionBatch.created_at.desc())
    result = await db.execute(query)
    return [BatchResponse.model_validate(b) for b in result.scalars()]


async def create_batch(db: AsyncSession, body: BatchCreate) -> BatchResponse:
    batch = ProductionBatch(
        batch_number=body.batch_number,
        product_id=uuid.UUID(body.product_id),
        planned_quantity=body.planned_quantity,
        start_date=body.start_date,
        end_date=body.end_date,
        notes=body.notes,
    )
    db.add(batch)
    await db.flush()
    return BatchResponse.model_validate(batch)


async def update_batch_status(
    db: AsyncSession,
    batch_id: str,
    status: str,
    produced_quantity: Optional[float] = None,
) -> BatchResponse:
    result = await db.execute(select(ProductionBatch).where(ProductionBatch.id == uuid.UUID(batch_id)))
    batch = result.scalar_one_or_none()
    if not batch:
        raise HTTPException(status_code=404, detail="Batch not found")
    batch.status = status
    if produced_quantity is not None:
        batch.produced_quantity = produced_quantity
    return BatchResponse.model_validate(batch)


async def list_platform_orders(
    db: AsyncSession,
    platform: Optional[str] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
) -> PaginatedResponse[PlatformOrderResponse]:
    query = select(PlatformOrder)
    if platform:
        query = query.where(PlatformOrder.platform == platform)
    if from_date:
        query = query.where(PlatformOrder.order_date >= from_date)
    if to_date:
        query = query.where(PlatformOrder.order_date <= to_date)
    if search:
        term = f"%{search}%"
        query = query.where(
            or_(PlatformOrder.product_name.ilike(term), PlatformOrder.order_id.ilike(term))
        )

    count_result = await db.execute(select(func.count()).select_from(query.subquery()))
    total = count_result.scalar_one()

    query = query.order_by(PlatformOrder.order_date.desc()).offset((page - 1) * limit).limit(limit)
    result = await db.execute(query)
    items = [PlatformOrderResponse.model_validate(o) for o in result.scalars()]
    return PaginatedResponse.build(items, total, page, limit)


async def create_platform_order(db: AsyncSession, body: PlatformOrderCreate) -> PlatformOrderResponse:
    order = PlatformOrder(
        platform=body.platform,
        order_id=body.order_id,
        product_id=uuid.UUID(body.product_id) if body.product_id else None,
        product_name=body.product_name,
        quantity=body.quantity,
        amount=body.amount,
        status=body.status,
        order_date=body.order_date,
    )
    db.add(order)
    await db.flush()
    return PlatformOrderResponse.model_validate(order)


async def update_order_status(
    db: AsyncSession,
    order_id: str,
    body,
) -> PlatformOrderResponse:
    from datetime import datetime, timezone
    result = await db.execute(select(PlatformOrder).where(PlatformOrder.id == uuid.UUID(order_id)))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    history = list(order.status_history or [])
    history.append({
        "status": body.status,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "note": body.note or "",
    })
    order.status = body.status
    order.status_history = history
    if body.tracking_number is not None:
        order.tracking_number = body.tracking_number
    if body.status == "returned":
        order.is_returned = True
        order.return_reason = body.note

    await db.flush()
    return PlatformOrderResponse.model_validate(order)


async def returns_analysis(db: AsyncSession) -> dict:
    result = await db.execute(select(PlatformOrder))
    orders = result.scalars().all()

    by_platform: dict[str, dict] = {}
    by_product: dict[str, dict] = {}

    for o in orders:
        p = o.platform
        prod = o.product_name
        by_platform.setdefault(p, {"total": 0, "returned": 0})
        by_product.setdefault(prod, {"total": 0, "returned": 0})

        by_platform[p]["total"] += 1
        by_product[prod]["total"] += 1

        if o.is_returned or o.status == "returned":
            by_platform[p]["returned"] += 1
            by_product[prod]["returned"] += 1

    platform_rows = [
        {
            "platform": k,
            "total_orders": v["total"],
            "returns": v["returned"],
            "return_rate": round(v["returned"] / v["total"] * 100, 1) if v["total"] else 0,
        }
        for k, v in sorted(by_platform.items())
    ]
    product_rows = [
        {
            "product": k,
            "total_orders": v["total"],
            "returns": v["returned"],
            "return_rate": round(v["returned"] / v["total"] * 100, 1) if v["total"] else 0,
        }
        for k, v in sorted(by_product.items(), key=lambda x: -x[1]["returned"])
    ]

    return {"by_platform": platform_rows, "by_product": product_rows}


async def platform_summary(
    db: AsyncSession, order_date: Optional[date] = None
) -> list[dict]:
    query = select(
        PlatformOrder.platform,
        func.count().label("count"),
        func.sum(PlatformOrder.amount).label("revenue"),
        func.sum(PlatformOrder.quantity).label("units"),
    ).group_by(PlatformOrder.platform)
    if order_date:
        query = query.where(PlatformOrder.order_date == order_date)
    result = await db.execute(query)
    return [
        {
            "platform": r.platform,
            "orders": r.count,
            "revenue": float(r.revenue or 0),
            "units": float(r.units or 0),
        }
        for r in result.all()
    ]


# ── Import / Export ───────────────────────────────────────────────────────────

IMPORT_COLUMNS = ["name", "sku", "category", "unit", "mrp", "cost_price", "gst_rate", "hsn_code", "reorder_level", "max_stock", "description"]


def get_import_template() -> StreamingResponse:
    """Return a sample Excel file the user can fill in and upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    ws.append(IMPORT_COLUMNS)
    for col_idx, col in enumerate(IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Sample row
    ws.append(["Backero Cream 50g", "BCR-001", "Skincare", "pcs", 299, 120, 18, "3304", 50, 500, "Herbal face cream"])

    # Column widths
    widths = [25, 15, 15, 8, 10, 12, 10, 12, 14, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Instruction sheet
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "INSTRUCTIONS"
    ws2["A1"].font = Font(bold=True, size=14)
    notes = [
        ("name", "Product name (required)"),
        ("sku", "Unique product code (required)"),
        ("category", "Product category (e.g. Skincare, Haircare)"),
        ("unit", "Unit of measure: pcs / kg / g / ml / L"),
        ("mrp", "MRP (selling price) in ₹"),
        ("cost_price", "Cost / purchase price in ₹"),
        ("gst_rate", "GST % (0, 5, 12, 18, or 28)"),
        ("hsn_code", "HSN/SAC code"),
        ("reorder_level", "Stock level to trigger reorder alert"),
        ("max_stock", "Maximum stock capacity"),
        ("description", "Optional product description"),
    ]
    for i, (col, note) in enumerate(notes, start=3):
        ws2.cell(row=i, column=1, value=col).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=note)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=product_import_template.xlsx"},
    )


async def export_products(db: AsyncSession) -> StreamingResponse:
    """Export all products as an Excel file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    result = await db.execute(
        select(Product).options(selectinload(Product.inventory)).where(Product.is_active == True).order_by(Product.name)
    )
    products = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["Name", "SKU", "Category", "Unit", "MRP (₹)", "Cost Price (₹)", "GST %", "HSN Code", "Current Stock", "Reorder Level", "Max Stock", "Low Stock", "Description"]
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    for p in products:
        inv = p.inventory
        ws.append([
            p.name, p.sku, p.category, p.unit,
            float(p.mrp), float(p.cost_price), float(p.gst_rate), p.hsn_code or "",
            float(inv.current_stock) if inv else 0,
            float(inv.reorder_level) if inv else 0,
            float(inv.max_stock) if inv else 0,
            "Yes" if (inv and inv.current_stock <= inv.reorder_level) else "No",
            p.description or "",
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=products_export.xlsx"},
    )


async def import_products(
    db: AsyncSession, file: UploadFile, actor_id: uuid.UUID, actor_name: str
) -> dict:
    """Import products from an uploaded Excel or CSV file."""
    import openpyxl

    content = await file.read()
    filename = (file.filename or "").lower()

    rows: list[dict] = []

    if filename.endswith(".csv"):
        import csv
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif filename.endswith((".xlsx", ".xls")):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers_row = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            rows.append(dict(zip(headers_row, row)))
    else:
        raise HTTPException(status_code=400, detail="Only .xlsx or .csv files are supported")

    created, skipped, errors = 0, 0, []

    for i, row in enumerate(rows, start=2):
        name = str(row.get("name") or "").strip()
        sku = str(row.get("sku") or "").strip()
        if not name or not sku:
            errors.append(f"Row {i}: name and sku are required")
            continue

        existing = await db.execute(select(Product).where(Product.sku == sku))
        if existing.scalar_one_or_none():
            skipped += 1
            continue

        def _float(v, default=0.0):
            try:
                return float(v) if v not in (None, "") else default
            except (ValueError, TypeError):
                return default

        product = Product(
            name=name,
            sku=sku,
            category=str(row.get("category") or "General").strip(),
            unit=str(row.get("unit") or "pcs").strip(),
            mrp=_float(row.get("mrp")),
            cost_price=_float(row.get("cost_price")),
            gst_rate=_float(row.get("gst_rate"), 18),
            hsn_code=str(row.get("hsn_code") or "").strip() or None,
            description=str(row.get("description") or "").strip() or None,
        )
        db.add(product)
        await db.flush()
        db.add(Inventory(
            product_id=product.id,
            current_stock=0,
            reorder_level=_float(row.get("reorder_level"), 10),
            max_stock=_float(row.get("max_stock"), 1000),
        ))
        created += 1

    from app.utils.activity import log as activity_log
    await activity_log(
        db,
        actor_id=actor_id,
        actor_name=actor_name,
        action="import",
        entity_type="product",
        entity_name=f"{created} products",
        description=f"{actor_name} imported {created} products from '{file.filename}' ({skipped} skipped, {len(errors)} errors)",
    )

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}
