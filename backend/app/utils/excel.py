import io
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Styling helpers ───────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header_row(ws, headers: list[str]):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 22


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 36)


def _to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(file_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(file_bytes))


def _parse_rows(ws) -> list[dict]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h or "").strip().lower().replace(" ", "_") for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


def _s(v, default="") -> str:
    return str(v).strip() if v is not None else default


def _f(v, default=0.0) -> float:
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


# ── Products ──────────────────────────────────────────────────────────────────

_PRODUCT_HEADERS = ["name", "sku", "category", "unit", "mrp", "cost_price", "gst_rate", "hsn_code", "reorder_level", "max_stock", "description"]
_PRODUCT_SAMPLE = [["Rose Face Cream", "RFC-001", "Skincare", "pcs", 299.0, 120.0, 18, "33049900", 50, 500, "Hydrating rose face cream"]]


def export_products(products: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    _header_row(ws, _PRODUCT_HEADERS)
    for p in products:
        ws.append([p.get(h) for h in _PRODUCT_HEADERS])
    _auto_width(ws)
    return _to_bytes(wb)


def products_sample() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    _header_row(ws, _PRODUCT_HEADERS)
    for row in _PRODUCT_SAMPLE:
        ws.append(row)
    _auto_width(ws)
    return _to_bytes(wb)


def parse_products(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    ws = _load(file_bytes).active
    errors: list[str] = []
    result: list[dict] = []
    for i, row in enumerate(_parse_rows(ws), start=2):
        name = _s(row.get("name"))
        sku = _s(row.get("sku"))
        if not name or not sku:
            errors.append(f"Row {i}: name and sku are required")
            continue
        result.append({
            "name": name, "sku": sku,
            "category": _s(row.get("category")) or None,
            "unit": _s(row.get("unit")) or "pcs",
            "mrp": _f(row.get("mrp")),
            "cost_price": _f(row.get("cost_price")),
            "gst_rate": _f(row.get("gst_rate"), 18),
            "hsn_code": _s(row.get("hsn_code")) or None,
            "reorder_level": _f(row.get("reorder_level"), 10),
            "max_stock": _f(row.get("max_stock"), 1000),
            "description": _s(row.get("description")) or None,
        })
    return result, errors


# ── Raw Materials ─────────────────────────────────────────────────────────────

_RM_HEADERS = ["name", "unit", "current_stock", "reorder_level", "cost_per_unit", "supplier", "notes"]
_RM_SAMPLE = [["Rose Oil", "kg", 25.0, 5.0, 850.0, "Kannauj Aromatic Co.", "Store in cool dark place"]]


def export_raw_materials(materials: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Materials"
    _header_row(ws, _RM_HEADERS)
    for m in materials:
        ws.append([m.get(h) for h in _RM_HEADERS])
    _auto_width(ws)
    return _to_bytes(wb)


def raw_materials_sample() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Raw Materials"
    _header_row(ws, _RM_HEADERS)
    for row in _RM_SAMPLE:
        ws.append(row)
    _auto_width(ws)
    return _to_bytes(wb)


def parse_raw_materials(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    ws = _load(file_bytes).active
    errors: list[str] = []
    result: list[dict] = []
    for i, row in enumerate(_parse_rows(ws), start=2):
        name = _s(row.get("name"))
        if not name:
            errors.append(f"Row {i}: name is required")
            continue
        result.append({
            "name": name,
            "unit": _s(row.get("unit")) or "kg",
            "current_stock": _f(row.get("current_stock")),
            "reorder_level": _f(row.get("reorder_level"), 5),
            "cost_per_unit": _f(row.get("cost_per_unit")),
            "supplier": _s(row.get("supplier")) or None,
            "notes": _s(row.get("notes")) or None,
        })
    return result, errors


# ── Platform Orders ───────────────────────────────────────────────────────────

_ORDER_HEADERS = ["platform", "order_id", "product_name", "quantity", "amount", "status", "order_date"]
_ORDER_SAMPLE = [["amazon", "AMZ-1001", "Rose Face Cream 50ml", 2, 598.0, "delivered", "2025-04-20"]]


def export_platform_orders(orders: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Platform Orders"
    _header_row(ws, _ORDER_HEADERS)
    for o in orders:
        ws.append([o.get(h) for h in _ORDER_HEADERS])
    _auto_width(ws)
    return _to_bytes(wb)


def platform_orders_sample() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Platform Orders"
    _header_row(ws, _ORDER_HEADERS)
    for row in _ORDER_SAMPLE:
        ws.append(row)
    _auto_width(ws)
    return _to_bytes(wb)


def parse_platform_orders(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    ws = _load(file_bytes).active
    errors: list[str] = []
    result: list[dict] = []
    VALID_PLATFORMS = {"amazon", "flipkart", "meesho", "website", "offline"}
    for i, row in enumerate(_parse_rows(ws), start=2):
        platform = _s(row.get("platform")).lower()
        order_id = _s(row.get("order_id"))
        product_name = _s(row.get("product_name"))
        order_date = _s(row.get("order_date"))
        if not platform or platform not in VALID_PLATFORMS:
            errors.append(f"Row {i}: platform must be one of {', '.join(VALID_PLATFORMS)}")
            continue
        if not order_id or not product_name or not order_date:
            errors.append(f"Row {i}: order_id, product_name, order_date are required")
            continue
        result.append({
            "platform": platform, "order_id": order_id,
            "product_name": product_name,
            "quantity": _f(row.get("quantity"), 1),
            "amount": _f(row.get("amount")),
            "status": _s(row.get("status")) or "delivered",
            "order_date": order_date,
        })
    return result, errors


# ── Finance Entries ───────────────────────────────────────────────────────────

_ENTRY_HEADERS = ["date", "type", "category", "description", "amount", "payment_mode", "reference"]
_ENTRY_SAMPLE = [
    ["2025-04-20", "income", "Sales", "Amazon settlement Apr W3", 42500.0, "bank", "UTR123456"],
    ["2025-04-21", "expense", "Packaging", "Box & label purchase", 8200.0, "upi", ""],
]


def export_entries(entries: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Entries"
    _header_row(ws, _ENTRY_HEADERS)
    for e in entries:
        ws.append([e.get(h) for h in _ENTRY_HEADERS])
    _auto_width(ws)
    return _to_bytes(wb)


def entries_sample() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Finance Entries"
    _header_row(ws, _ENTRY_HEADERS)
    for row in _ENTRY_SAMPLE:
        ws.append(row)
    _auto_width(ws)
    return _to_bytes(wb)


def parse_entries(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    ws = _load(file_bytes).active
    errors: list[str] = []
    result: list[dict] = []
    VALID_TYPES = {"income", "expense"}
    VALID_MODES = {"cash", "bank", "upi", "cheque"}
    for i, row in enumerate(_parse_rows(ws), start=2):
        row_date = _s(row.get("date"))
        row_type = _s(row.get("type")).lower()
        category = _s(row.get("category"))
        description = _s(row.get("description"))
        if not row_date:
            errors.append(f"Row {i}: date is required")
            continue
        if row_type not in VALID_TYPES:
            errors.append(f"Row {i}: type must be income or expense")
            continue
        if not category or not description:
            errors.append(f"Row {i}: category and description are required")
            continue
        mode = _s(row.get("payment_mode")).lower() or "cash"
        if mode not in VALID_MODES:
            mode = "cash"
        result.append({
            "date": row_date, "type": row_type,
            "category": category, "description": description,
            "amount": _f(row.get("amount")),
            "payment_mode": mode,
            "reference": _s(row.get("reference")) or None,
        })
    return result, errors


# ── Invoices export-only ──────────────────────────────────────────────────────

def export_invoices(invoices: list[dict]) -> bytes:
    headers = ["invoice_number", "invoice_date", "due_date", "customer_name", "customer_phone",
               "customer_gstin", "subtotal", "cgst", "sgst", "total", "is_gst", "status"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"
    _header_row(ws, headers)
    for inv in invoices:
        ws.append([inv.get(h) for h in headers])
    _auto_width(ws)
    return _to_bytes(wb)


# ── Employees ─────────────────────────────────────────────────────────────────

_EMP_HEADERS = ["name", "phone", "email", "designation", "department_name", "salary", "join_date"]
_EMP_SAMPLE = [["Priya Sharma", "9876543210", "priya@backero.in", "Sales Executive", "Sales", 28000, "2024-01-15"]]


def export_employees(employees: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    _header_row(ws, _EMP_HEADERS)
    for e in employees:
        ws.append([e.get(h) for h in _EMP_HEADERS])
    _auto_width(ws)
    return _to_bytes(wb)


def employees_sample() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    _header_row(ws, _EMP_HEADERS)
    for row in _EMP_SAMPLE:
        ws.append(row)
    _auto_width(ws)
    return _to_bytes(wb)


def parse_employees(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    ws = _load(file_bytes).active
    errors: list[str] = []
    result: list[dict] = []
    for i, row in enumerate(_parse_rows(ws), start=2):
        name = _s(row.get("name"))
        phone = _s(row.get("phone"))
        if not name or not phone:
            errors.append(f"Row {i}: name and phone are required")
            continue
        result.append({
            "name": name, "phone": phone,
            "email": _s(row.get("email")) or None,
            "designation": _s(row.get("designation")) or None,
            "salary": _f(row.get("salary")) or None,
            "join_date": _s(row.get("join_date")) or None,
        })
    return result, errors


# ── Tasks export-only ─────────────────────────────────────────────────────────

def export_tasks(tasks: list[dict]) -> bytes:
    headers = ["title", "priority", "status", "due_date", "assigned_to", "created_by", "description"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    _header_row(ws, headers)
    for t in tasks:
        ws.append([
            t.get("title"), t.get("priority"), t.get("status"),
            t.get("due_date"), t.get("assigned_to"), t.get("created_by"),
            t.get("description"),
        ])
    _auto_width(ws)
    return _to_bytes(wb)
